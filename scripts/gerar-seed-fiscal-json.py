# -*- coding: utf-8 -*-
"""
Converte as planilhas oficiais de NCM, CFOP e CEST (arquivos .xlsx na raiz do
projeto) para os arquivos JSON de fallback local usados por
src/modules/fiscal/seed-fiscal/fonte-externa.service.ts quando a variável de
ambiente SEED_FISCAL_<TABELA>_URL não está configurada.

Uso: python scripts/gerar-seed-fiscal-json.py
"""
import json
import re
import openpyxl

OUT_DIR = 'src/modules/fiscal/seed-fiscal/data'

# ---------------------------------------------------------------------------
# CFOP — 160314_Tabela_CFOP.xlsx (planilha "CFOP", colunas: CFOP, Descrição
# Resumida, indNFe, indComunica, indTransp, indDevol)
# ---------------------------------------------------------------------------
def gerar_cfop():
    wb = openpyxl.load_workbook('160314_Tabela_CFOP.xlsx', data_only=True)
    ws = wb['CFOP']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # CFOP 5151 vem sem descrição na planilha oficial; corrigido por analogia
    # ao CFOP 6151 (mesma operação, âmbito interestadual): "Transferência de
    # produção do estabelecimento".
    correcao_descricao = {5151: 'Transferência de produção do estabelecimento'}

    def derivar(codigo):
        primeiro = str(codigo)[0]
        tipo = 'ENTRADA' if primeiro in ('1', '2', '3') else 'SAIDA'
        mapa_ambito = {'1': 'ESTADUAL', '5': 'ESTADUAL', '2': 'INTERESTADUAL',
                       '6': 'INTERESTADUAL', '3': 'EXTERIOR', '7': 'EXTERIOR'}
        return tipo, mapa_ambito.get(primeiro, 'ESTADUAL')

    registros = []
    for cfop, desc_resumida, _indNFe, _indComunica, _indTransp, _indDevol in rows:
        codigo = str(cfop)
        descricao = desc_resumida if desc_resumida is not None else correcao_descricao.get(cfop)
        if descricao is None:
            raise ValueError(f'Sem descricao para CFOP {codigo}')
        descricao = descricao.strip()
        tipo, ambito = derivar(cfop)
        registros.append({
            'codigo': codigo,
            'descricao': descricao,
            'tipo': tipo,
            'ambito': ambito,
            'geraCredIcms': False,
            'geraCredPisCofins': False,
            'incideIpi': False,
        })

    with open(f'{OUT_DIR}/cfop.json', 'w', encoding='utf-8') as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)
    print(f'cfop.json: {len(registros)} registros')


# ---------------------------------------------------------------------------
# NCM — Tabela_NCM_Vigente_20260711.xlsx (planilha "Tabela NCM", cabeçalho na
# linha 5: Código, Descrição, Data Início, Data Fim, Ato Legal Início,
# Número, Ano). Apenas os códigos "folha" (8 dígitos, formato NNNN.NN.NN) são
# NCMs completos — os demais são capítulos/posições/subposições
# intermediárias da hierarquia, não usados como código de produto.
# ---------------------------------------------------------------------------
def gerar_ncm():
    wb = openpyxl.load_workbook('Tabela_NCM_Vigente_20260711.xlsx', data_only=True)
    ws = wb['Tabela NCM']
    rows = list(ws.iter_rows(min_row=6, values_only=True))

    padrao_folha = re.compile(r'^\d{4}\.\d{2}\.\d{2}$')
    registros = []
    for row in rows:
        codigo_raw = row[0]
        descricao = row[1]
        if not codigo_raw or not padrao_folha.match(str(codigo_raw)):
            continue
        codigo = str(codigo_raw).replace('.', '')
        if not descricao:
            continue
        descricao = descricao.strip()[:500]
        registros.append({
            'codigo': codigo,
            'descricao': descricao,
        })

    with open(f'{OUT_DIR}/ncm.json', 'w', encoding='utf-8') as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)
    print(f'ncm.json: {len(registros)} registros')


# ---------------------------------------------------------------------------
# CEST — Tabela_CEST_extraida.xlsx (texto extraído do PDF oficial do Convênio
# ICMS nº 142/2018 — Anexos II a XXVI, cada um trazendo os itens CEST de um
# segmento de mercadoria; Anexo I traz a lista Segmento -> Código de Segmento
# usada aqui só para referência, já que o próprio código CEST (2 primeiros
# dígitos) já identifica o segmento).
#
# Cada item aparece como uma linha "<item> <CEST> <NCM> <início da
# descrição>", seguida por 0+ linhas de continuação: linhas só com
# dígitos/pontos são fragmentos de NCM adicionais (SKU já tem múltiplos NCMs
# por CEST) e são ignoradas; linhas começando com um fragmento de NCM seguido
# de texto têm o NCM removido e o texto reaproveitado como continuação da
# descrição; as demais linhas são concatenadas diretamente à descrição.
#
# Os Anexos XXVII (bens não sujeitos a ST fabricados em escala industrial não
# relevante), XXVIII (formulário de credenciamento) e XXIX (relação de
# contribuintes credenciados) não contêm itens CEST no formato-alvo desta
# tabela e são ignorados.
# ---------------------------------------------------------------------------
def gerar_cest():
    wb = openpyxl.load_workbook('Tabela_CEST_extraida.xlsx', data_only=True)
    ws = wb['Sheet']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    texts = [r[1] for r in rows if r[1]]

    # Anexo I: "<item> <nome do segmento> <código do segmento>" — mapeia o
    # código de 2 dígitos (mesmo prefixo dos 2 primeiros dígitos do CEST) ao
    # nome oficial do segmento de mercadoria. Restrito ao intervalo do
    # próprio Anexo I para não casar falsamente com linhas de outros anexos.
    segmento_re = re.compile(r'^(\d{2})\s+(.*\S)\s+(\d{2})$')
    anexo_re_preview = re.compile(r'^ANEXO\s+([IVXL]+)\b')
    mapa_segmento = {}
    dentro_anexo_i = False
    for texto in texts:
        texto_normalizado_preview = texto.strip()
        anexo_match_preview = anexo_re_preview.match(texto_normalizado_preview)
        if anexo_match_preview:
            dentro_anexo_i = anexo_match_preview.group(1) == 'I'
            continue
        if not dentro_anexo_i:
            continue
        m = segmento_re.match(texto_normalizado_preview)
        if m:
            mapa_segmento[m.group(3).zfill(2)] = m.group(2).strip()

    item_re = re.compile(r'^(\d+(?:\.\d+)?)\s+(\d{2}\.\d{3}\.\d{2})\s+(\S+)\s+(.*)$')
    ncm_only_re = re.compile(r'^[\d\.\s]+$')
    ncm_frag_re = re.compile(r'^([\d\.]{4,})\s+(.*)$')
    anexo_re = re.compile(r'^ANEXO\s+([IVXL]+)\b')
    anexos_excluidos = {'I', 'XXVII', 'XXVIII', 'XXIX'}

    # Um pequeno grupo de itens do Anexo XXVI (55.0 a 64.0) usa uma lista de
    # "Capítulo(s)" da NCM em vez de um código NCM/SH numérico único, e essa
    # lista continua nas linhas seguintes junto com o início real da
    # descrição na mesma linha física do PDF (ex.: "34 itens deste anexo" =
    # capítulo "34" + descrição "itens deste anexo"). Essa função remove os
    # tokens iniciais puramente numéricos (com vírgula opcional) e a
    # conjunção "e" entre eles, retornando só o texto restante — que pode
    # ficar vazio quando a linha inteira é apenas continuação da lista.
    def remover_fragmento_lista_capitulos(texto):
        """Remove tokens iniciais que são números (com vírgula opcional) ou
        as conjunções 'e'/'a' entre eles (ex.: "52, 55, 58, 63", "e 65",
        "15 a 23"), retornando o texto restante (pode ficar vazio)."""
        tokens = texto.split()
        i = 0
        while i < len(tokens):
            token_sem_virgula = tokens[i].rstrip(',')
            if token_sem_virgula.isdigit() or tokens[i] in ('e', 'a'):
                i += 1
                continue
            break
        return ' '.join(tokens[i:])

    registros = []
    atual = None
    dentro_do_intervalo_valido = False

    def concluir_registro():
        nonlocal atual
        if atual:
            descricao = ' '.join(p for p in atual['descricao_partes'] if p).strip()[:500]
            # Modelo Cest.codigo é VarChar(7) e a validação do seed exige
            # exatamente 7 dígitos sem pontuação (mesmo padrão de NCM/CFOP).
            codigo_sem_pontos = atual['codigo'].replace('.', '')
            segmento = mapa_segmento.get(codigo_sem_pontos[:2])
            registro = {'codigo': codigo_sem_pontos, 'descricao': descricao}
            if segmento:
                registro['segmento'] = segmento
            registros.append(registro)
        atual = None

    for texto in texts:
        texto_normalizado = texto.strip()

        anexo_match = anexo_re.match(texto_normalizado)
        if anexo_match:
            concluir_registro()
            dentro_do_intervalo_valido = anexo_match.group(1) not in anexos_excluidos
            continue

        if not dentro_do_intervalo_valido:
            continue

        # Cabeçalhos de coluna repetidos em cada página do PDF de origem.
        if texto_normalizado.upper().startswith('ITEM CEST') or texto_normalizado.upper().startswith('ITEM NOME'):
            continue

        item_match = item_re.match(texto)
        if item_match:
            concluir_registro()
            campo_ncm = item_match.group(3)
            descricao_inicial = item_match.group(4).strip()
            # Remove pontuação de fim de campo (vírgula ou ponto) antes de
            # testar se é um NCM/SH numérico — alguns itens (ex.: 25.020.00
            # do Anexo XXIV) têm o primeiro NCM de uma lista seguido de
            # vírgula na mesma linha, indicando que a lista continua na(s)
            # linha(s) seguinte(s) (tratadas por ncm_frag_re/ncm_only_re).
            campo_ncm_normalizado = campo_ncm.rstrip(',.').replace('.', '')
            if campo_ncm_normalizado.isdigit():
                # Caso comum: campo_ncm é de fato um código NCM/SH numérico.
                atual = {'codigo': item_match.group(2), 'descricao_partes': [descricao_inicial], 'lista_capitulos': False}
            elif re.match(r'^cap[ií]tulos?$', campo_ncm, re.IGNORECASE):
                # "Capítulo(s) NN, NN ..." — a lista de capítulos continua
                # nas linhas seguintes; remove os números da lista já
                # presentes nesta primeira linha antes de guardar a descrição.
                atual = {'codigo': item_match.group(2), 'descricao_partes': [remover_fragmento_lista_capitulos(descricao_inicial)], 'lista_capitulos': True}
            else:
                # Sem NCM/SH informado — o "campo_ncm" capturado é na
                # verdade a primeira palavra da própria descrição.
                atual = {'codigo': item_match.group(2), 'descricao_partes': [f'{campo_ncm} {descricao_inicial}'.strip()], 'lista_capitulos': False}
            continue

        if atual is None:
            # Linha fora de um item reconhecido (título de segmento, nota de
            # rodapé, etc.) — não pertence à descrição de nenhum item.
            continue

        if ncm_only_re.match(texto):
            # Fragmento de NCM adicional do mesmo CEST, sem texto de descrição.
            continue

        ncm_frag_match = ncm_frag_re.match(texto)
        if ncm_frag_match and not atual['lista_capitulos']:
            atual['descricao_partes'].append(ncm_frag_match.group(2).strip())
            continue

        if atual['lista_capitulos']:
            atual['descricao_partes'].append(remover_fragmento_lista_capitulos(texto_normalizado))
            continue

        atual['descricao_partes'].append(texto_normalizado)

    concluir_registro()

    with open(f'{OUT_DIR}/cest.json', 'w', encoding='utf-8') as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)
    print(f'cest.json: {len(registros)} registros')


if __name__ == '__main__':
    gerar_cfop()
    gerar_ncm()
    gerar_cest()
